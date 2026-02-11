# products/views_receipt.py - FIXED VERSION
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, time, timedelta

from django.http import FileResponse, JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.views import View
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator

from django.utils.dateparse import parse_date
from django.db.models import Q

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status

from reportlab.pdfgen import canvas
from reportlab.graphics.barcode import code128
from reportlab.lib.pagesizes import A4, landscape

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Sale, SaleLine, SalePayment, CreditNote


TWOPLACES = Decimal("0.01")


def q2(x):
    return (Decimal(x or 0)).quantize(TWOPLACES, rounding=ROUND_HALF_UP)


def _user_location_code(user):
    emp = getattr(user, "employee", None)
    outlet = getattr(emp, "outlet", None) if emp else None
    loc = getattr(outlet, "location", None) if outlet else None
    return (getattr(loc, "code", "") or "").strip().upper()


def _user_location_name(user):
    emp = getattr(user, "employee", None)
    outlet = getattr(emp, "outlet", None) if emp else None
    loc = getattr(outlet, "location", None) if outlet else None
    name = (getattr(loc, "name", "") or "").strip()
    code = (getattr(loc, "code", "") or "").strip()
    return name or code or ""


def _can_view_sale(user, sale: Sale) -> bool:
    # Admin/superuser can view everything
    if getattr(user, "is_superuser", False) or getattr(user, "is_staff", False):
        return True

    # Manager/staff should only see their location (best-effort)
    user_loc = _user_location_code(user)
    if not user_loc:
        return True  # if user has no outlet mapping, don't block

    created_by = getattr(sale, "created_by", None)
    emp = getattr(created_by, "employee", None) if created_by else None
    outlet = getattr(emp, "outlet", None) if emp else None
    loc = getattr(outlet, "location", None) if outlet else None
    sale_loc = (getattr(loc, "code", "") or "").strip().upper()

    # if sale has no created_by/location info, allow
    if not sale_loc:
        return True

    return sale_loc == user_loc


def _inr_words(amount: Decimal) -> str:
    n = int(q2(amount))
    if n == 0:
        return "Rupees Zero Only"

    ones = [
        "", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
        "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen",
        "Sixteen", "Seventeen", "Eighteen", "Nineteen",
    ]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def two(num):
        if num < 20:
            return ones[num]
        return (tens[num // 10] + (" " + ones[num % 10] if num % 10 else "")).strip()

    def three(num):
        h = num // 100
        r = num % 100
        if h and r:
            return f"{ones[h]} Hundred {two(r)}".strip()
        if h:
            return f"{ones[h]} Hundred".strip()
        return two(r).strip()

    parts = []
    crore = n // 10000000
    n %= 10000000
    lakh = n // 100000
    n %= 100000
    thousand = n // 1000
    n %= 1000
    hundred = n

    if crore:
        parts.append(f"{three(crore)} Crore")
    if lakh:
        parts.append(f"{three(lakh)} Lakh")
    if thousand:
        parts.append(f"{three(thousand)} Thousand")
    if hundred:
        parts.append(three(hundred))

    return "Rupees " + " ".join([p for p in parts if p]).strip() + " Only"


@method_decorator(login_required, name="dispatch")
class SaleReceiptPdfView(View):
    def get(self, request, invoice_no: str):
        sale = get_object_or_404(Sale, invoice_no=invoice_no)

        if not _can_view_sale(request.user, sale):
            return JsonResponse({"detail": "Not allowed."}, status=403)

        loc_code = _user_location_code(request.user)
        loc_name = _user_location_name(request.user)

        if loc_code == "UV":
            shop_at = "SHOP AT"
            shop_domain = "Brands4Less"
            gstin = "06AADFW9945P1ZB"
            address_line = "268, Rao Gajraj Road, Phase IV, Gurugram 122015"
            place_supply = "Haryana"
            customer_address = "Gurugram"
        else:
            shop_at = "SHOP AT"
            shop_domain = ""
            gstin = ""
            address_line = ""
            place_supply = ""
            customer_address = ""

        now = timezone.localtime(timezone.now())
        date_str = now.strftime("%d/%m/%Y")
        time_str = now.strftime("%I:%M:%S %p")

        cust = getattr(sale, "customer", None)
        cust_name = (getattr(cust, "name", "") or "").strip()
        cust_phone = (getattr(cust, "phone", "") or "").strip()

        salesman = getattr(sale, "salesman", None)
        salesman_name = ""
        if salesman:
            u = getattr(salesman, "user", None)
            if u:
                salesman_name = (u.get_full_name() or "").strip() or (u.username or "")
            else:
                salesman_name = (getattr(salesman, "name", "") or "").strip()

        lines = list(SaleLine.objects.filter(sale=sale).order_by("id"))

        total_qty = Decimal("0.00")
        total_disc = Decimal("0.00")
        total_net = Decimal("0.00")

        # ------------------ BILL DISCOUNT FALLBACK ------------------
        # If UI saved discount at Sale header (sale.discount_total) but not per-line,
        # distribute it across lines proportionally to line gross.
        bill_disc = q2(getattr(sale, "discount_total", 0) or 0)

        def _line_has_any_discount(ln: SaleLine) -> bool:
            dp = q2(getattr(ln, "discount_percent", 0) or 0)
            da = q2(getattr(ln, "discount_amount", 0) or 0)
            return (dp > 0) or (da > 0)

        allocate_bill_discount = (
            bill_disc > 0
            and lines
            and (not any(_line_has_any_discount(ln) for ln in lines))
        )

        bill_disc_by_line_id = {}
        if allocate_bill_discount:
            gross_list = []
            gross_total = Decimal("0.00")
            for ln in lines:
                qty = Decimal(int(getattr(ln, "qty", 0) or 0)) or Decimal("1")
                sp = q2(getattr(ln, "sp", 0) or 0)
                g = q2(sp * qty)
                gross_list.append((ln.id, g, qty, sp))
                gross_total = q2(gross_total + g)

            if gross_total > 0:
                running = Decimal("0.00")
                # allocate with rounding, fix last line to match exact bill_disc
                for i, (lid, g, qty, sp) in enumerate(gross_list):
                    if i == len(gross_list) - 1:
                        d = q2(bill_disc - running)
                    else:
                        d = (bill_disc * g / gross_total).quantize(
                            TWOPLACES, rounding=ROUND_HALF_UP
                        )
                        running = q2(running + d)
                    # safety clamp
                    if d < 0:
                        d = Decimal("0.00")
                    if d > g:
                        d = g
                    bill_disc_by_line_id[lid] = q2(d)

        def per_unit_disc(ln: SaleLine) -> Decimal:
            # ✅ If bill discount was allocated, use it.
            if ln.id in bill_disc_by_line_id:
                qty = Decimal(int(getattr(ln, "qty", 0) or 0)) or Decimal("1")
                return q2(bill_disc_by_line_id[ln.id] / qty)

            # ✅ Discount is calculated on SP
            sp = q2(getattr(ln, "sp", 0) or 0)
            dp = q2(getattr(ln, "discount_percent", 0) or 0)
            da = q2(getattr(ln, "discount_amount", 0) or 0)  # per-unit stored

            # ✅ Fallback: if discount fields are empty but unit_cost is saved as net
            unit_cost = q2(getattr(ln, "unit_cost", None) or 0)
            if dp <= 0 and da <= 0 and unit_cost > 0 and unit_cost < sp:
                d = q2(sp - unit_cost)
            elif dp > 0:
                d = (sp * dp / Decimal("100")).quantize(TWOPLACES, rounding=ROUND_HALF_UP)
            else:
                d = da

            if d < 0:
                d = Decimal("0.00")
            if d > sp:
                d = sp
            return q2(d)

        # ---------------- THERMAL RECEIPT PAGE (prints like machine) ----------------
        # 80mm roll width -> ~226.77 points (72pt/in). Keep small margins.
        PAGE_W = 226.8
        M = 8.0

        # Dynamic height so printer doesn't scale to A4 (main reason your print was blank/huge).
        # Height estimate in points; we add enough space for lines + footer.
        est_h = 520 + (len(lines) * 14)
        PAGE_H = max(600, est_h)

        buf = BytesIO()
        c = canvas.Canvas(buf, pagesize=(PAGE_W, PAGE_H))

        y = PAGE_H - 18

        def txt(x, y, s, size=8, bold=False):
            c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
            c.drawString(x, y, s or "")

        def rtxt(x, y, s, size=8, bold=False):
            c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
            c.drawRightString(x, y, s or "")

        def center(y, s, size=9, bold=False):
            c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
            c.drawCentredString(PAGE_W / 2, y, s or "")

        def dashline(y):
            c.saveState()
            c.setDash(2, 2)
            c.line(M, y, PAGE_W - M, y)
            c.restoreState()

        def _tax_rate_for_total(total: Decimal) -> Decimal:
            # <= 2500 => 5%, else 18%
            return Decimal("0.05") if q2(total) <= Decimal("2500.00") else Decimal("0.18")

        def _calc_tax_summary(total: Decimal):
            """
            Rules (corrected):
            - rate = 5% if total <= 2500 else 18%
            - taxable_value = total - (total * rate)
            - cgst = taxable_value * rate
            - sgst = taxable_value * rate
            - cess, igst = 0.00
            """
            rate = _tax_rate_for_total(total)
            taxable_value = q2(total - (total * rate))
            cgst = q2(taxable_value * rate)
            sgst = q2(taxable_value * rate)
            cess = Decimal("0.00")
            igst = Decimal("0.00")
            return taxable_value, cgst, sgst, cess, igst

        # Header (no "DUPLICATE")
        center(y, (loc_code or "OUTLET")[:12], 9, True)
        y -= 14
        center(y, "Receipt", 10, True)
        y -= 12

        if address_line:
            center(y, address_line, 7, False)
            y -= 10

        if gstin:
            center(y, f"GSTIN NO : {gstin}", 7, False)
        else:
            center(y, "GSTIN NO : ", 7, False)
        y -= 10

        # Shop at (keep same placement style, just fit roll width)
        if shop_domain:
            center(y, f"{shop_at}", 7, True)
            y -= 10
            center(y, f"{shop_domain}", 8, True)
            y -= 9
            center(y, "A unit of Wabi-Sabi", 7, False)
            y -= 10

        # Customer + meta
        txt(M, y, f"Name   : {cust_name}", 8, False)
        y -= 11
        txt(M, y, f"Mobile : {cust_phone}", 8, False)
        y -= 11

        # Receipt No (not Credit Note)
        txt(M, y, f"Receipt No : {sale.invoice_no}", 8, False)
        y -= 11

        txt(M, y, f"Date : {date_str}", 7, False)
        y -= 10
        txt(M, y, f"Time : {time_str}", 7, False)
        y -= 10
        txt(M, y, f"Salesman : {salesman_name}", 7, False)
        y -= 10

        dashline(y)
        y -= 12

        # Table header (reduced gap so right columns sit under the horizontal line)
        # Columns tuned for 80mm:
        # Item (left), Qty, SP, Disc, Net (right aligned)
        item_x = M
        qty_x = 118
        sp_x = 148
        disc_x = 178
        net_x = PAGE_W - M

        txt(item_x, y, "Item", 8, True)
        rtxt(qty_x, y, "Qty", 8, True)
        rtxt(sp_x, y, "SP", 8, True)      # MRP -> SP
        rtxt(disc_x, y, "Disc", 8, True)
        rtxt(net_x, y, "Net", 8, True)
        y -= 9

        dashline(y)
        y -= 12

        # Lines
        for ln in lines:
            qty = Decimal(int(getattr(ln, "qty", 0) or 0))
            sp = q2(getattr(ln, "sp", 0) or 0)

            dpu = per_unit_disc(ln)
            disc_line = q2(dpu * qty)

            # ✅ if bill-level discount is allocated, ensure disc_line matches allocation exactly
            if ln.id in bill_disc_by_line_id:
                disc_line = q2(bill_disc_by_line_id[ln.id])

            net_unit = q2(sp - dpu)
            net_line = q2(net_unit * qty)

            total_qty += qty
            total_disc += disc_line
            total_net += net_line

            pname = ""
            prod = getattr(ln, "product", None)
            ti = getattr(prod, "task_item", None) if prod else None
            if ti:
                pname = (
                    getattr(ti, "item_print_friendly_name", None)
                    or getattr(ti, "item_vasy_name", None)
                    or getattr(ti, "item_full_name", None)
                    or ""
                )
            if not pname:
                pname = (getattr(ln, "barcode", "") or "")

            # 2-line item name if long (thermal friendly)
            name = (pname or "").strip()
            if len(name) > 22:
                txt(item_x, y, name[:22], 7, False)
                y -= 10
                txt(item_x, y, name[22:44], 7, False)
            else:
                txt(item_x, y, name, 7, False)

            rtxt(qty_x, y, f"{qty:.1f}", 7, False)
            rtxt(sp_x, y, f"{sp:.2f}", 7, False)
            rtxt(disc_x, y, f"{disc_line:.2f}", 7, False)
            rtxt(net_x, y, f"{net_line:.2f}", 7, False)

            y -= 14

            # If height ever goes too low, extend page (avoid cutting on roll)
            if y < 140:
                c.showPage()
                c.setPageSize((PAGE_W, PAGE_H))
                y = PAGE_H - 18

        dashline(y)
        y -= 14

        txt(M, y, "TOTAL", 9, True)
        rtxt(net_x, y, f"{total_net:.2f}", 9, True)
        y -= 12

        txt(M, y, "ROUND OFF", 8, False)
        rtxt(net_x, y, "0.00", 8, False)
        y -= 12

        dashline(y)
        y -= 14

        txt(M, y, f"NO OF QTY : {total_qty:.2f}", 8, False)
        y -= 16

        txt(M, y, f"You Saved Rs. : {total_disc:.2f}", 9, True)
        y -= 14

        txt(M, y, _inr_words(total_net), 7, False)
        y -= 14

        txt(M, y, "Prices are inclusive of all taxes - Place of Supply :", 6, False)
        y -= 9
        txt(M, y, place_supply, 6, False)
        y -= 14

        # TAX SUMMARY (table like your image)
        center(y, "TAX SUMMARY", 8, True)
        y -= 10

        taxable_value, cgst, sgst, cess, igst = _calc_tax_summary(total_net)

        # Table geometry (fit roll width)
        table_left = M
        table_right = PAGE_W - M
        table_w = table_right - table_left

        cols = 5
        col_w = table_w / cols
        table_top = y
        row_h1 = 18
        row_h2 = 16

        # Draw outer box
        c.rect(table_left, table_top - (row_h1 + row_h2), table_w, (row_h1 + row_h2), stroke=1, fill=0)

        # Vertical lines
        for i in range(1, cols):
            x = table_left + (col_w * i)
            c.line(x, table_top, x, table_top - (row_h1 + row_h2))

        # Horizontal split between header and values
        c.line(table_left, table_top - row_h1, table_right, table_top - row_h1)

        # Header labels
        headers = ["TAXABLE\nVALUE", "CGST", "SGST", "Cess", "IGST"]
        for i, htxt in enumerate(headers):
            cx = table_left + col_w * i + col_w / 2
            c.setFont("Helvetica-Bold", 7)
            if "\n" in htxt:
                a, b = htxt.split("\n", 1)
                c.drawCentredString(cx, table_top - 10, a)
                c.drawCentredString(cx, table_top - 18, b)
            else:
                c.drawCentredString(cx, table_top - 14, htxt)

        # Values
        vals = [
            f"{taxable_value:.2f}",
            f"{cgst:.2f}",
            f"{sgst:.2f}",
            f"{cess:.2f}",
            f"{igst:.2f}",
        ]
        c.setFont("Helvetica", 7)
        for i, v in enumerate(vals):
            cx = table_left + col_w * i + col_w / 2
            c.drawCentredString(cx, table_top - row_h1 - 12, v)

        y = table_top - (row_h1 + row_h2) - 12

        txt(M, y, "Customer Details :", 8, False)
        y -= 10

        if customer_address:
            txt(M, y, f"Address : {customer_address}", 8, False)
        else:
            txt(M, y, "Address : ", 8, False)

        # ✅ extra space between Gurugram and barcode
        y -= 40

        # Barcode (centered on roll)
        bc_val = sale.invoice_no
        barcode_obj = code128.Code128(bc_val, barHeight=34, barWidth=0.8)
        bw = barcode_obj.width
        barcode_obj.drawOn(c, (PAGE_W - bw) / 2, y)

        y -= 48
        center(y, "Thank you for shopping with us", 8, True)
        y -= 12
        center(y, "For Home Delivery", 8, False)
        y -= 12
        center(y, "7303467070", 10, True)
        y -= 16

        printed = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %I:%M:%S %p")
        txt(M, y, f"Printed On : {printed}", 7, False)
        y -= 10
        txt(M, y, "Printed By : ", 7, False)

        c.showPage()
        c.save()

        buf.seek(0)
        return FileResponse(
            buf,
            as_attachment=False,
            filename=f"{sale.invoice_no}.pdf",
            content_type="application/pdf",
        )


# -------------------- Daywise Sales Summary helpers --------------------

def _is_admin(user) -> bool:
    return bool(getattr(user, "is_superuser", False) or getattr(user, "is_staff", False))


def _per_unit_discount(ln: SaleLine) -> Decimal:
    """
    Per-unit discount based on SP:
    - if discount_percent > 0 => SP * % / 100
    - else use discount_amount (stored per unit)
    - fallback: infer from unit_cost if discount fields are 0
    """
    sp = q2(getattr(ln, "sp", 0) or 0)
    dp = q2(getattr(ln, "discount_percent", 0) or 0)
    da = q2(getattr(ln, "discount_amount", 0) or 0)
    unit_cost = q2(getattr(ln, "unit_cost", None) or 0)

    if dp <= 0 and da <= 0 and unit_cost > 0 and unit_cost < sp:
        d = q2(sp - unit_cost)
    elif dp > 0:
        d = (sp * dp / Decimal("100")).quantize(TWOPLACES, rounding=ROUND_HALF_UP)
    else:
        d = da

    if d < 0:
        d = Decimal("0.00")
    if d > sp:
        d = sp
    return q2(d)


def _rate_for_unit_amount(unit_amount: Decimal) -> Decimal:
    """
    Tax rule:
      - <= 2500 => 5%
      - > 2500  => 18%
    """
    return Decimal("0.05") if q2(unit_amount) <= Decimal("2500.00") else Decimal("0.18")


def _split_inclusive_amount(inclusive_total: Decimal, rate: Decimal):
    """
    inclusive_total includes tax.
    taxable = inclusive_total / (1 + rate)
    tax     = inclusive_total - taxable
    cgst/sgst = tax/2
    """
    inclusive_total = q2(inclusive_total)
    if inclusive_total <= 0:
        z = Decimal("0.00")
        return z, z, z, z

    denom = (Decimal("1.00") + q2(rate))
    taxable = (inclusive_total / denom).quantize(TWOPLACES, rounding=ROUND_HALF_UP)
    tax = q2(inclusive_total - taxable)
    half = (tax / Decimal("2.00")).quantize(TWOPLACES, rounding=ROUND_HALF_UP)
    # keep totals consistent (cgst+sgst == tax)
    cgst = half
    sgst = q2(tax - cgst)
    return taxable, tax, cgst, sgst


class DaywiseSalesSummary(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """
        Query:
          - date_from=YYYY-MM-DD
          - date_to=YYYY-MM-DD
          - location=... (can repeat multiple times)
          - export=pdf | excel
        Response:
          { rows: [...], totals: {...} }
        """

        # -------- Parse dates --------
        df = parse_date((request.query_params.get("date_from") or "").strip())
        dt = parse_date((request.query_params.get("date_to") or "").strip())

        if not df or not dt:
            return Response(
                {"detail": "date_from and date_to are required (YYYY-MM-DD)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if dt < df:
            df, dt = dt, df

        tz = timezone.get_current_timezone()
        start_dt = timezone.make_aware(datetime.combine(df, time.min), tz)
        end_dt = timezone.make_aware(datetime.combine(dt, time.max), tz)

        # -------- Location filtering --------
        loc_params = request.query_params.getlist("location")
        loc_params = [str(x).strip() for x in loc_params if str(x).strip()]

        # non-admin must be restricted to their own outlet
        if not _is_admin(request.user):
            user_loc = _user_location_code(request.user)
            if user_loc:
                loc_params = [user_loc]

        # -------- Fetch sales in range --------
        sales_qs = Sale.objects.filter(
            transaction_date__range=(start_dt, end_dt)
        ).order_by("transaction_date", "id")

        # ✅ FIXED: Improved location filtering with proper Q object usage and distinct()
        if loc_params:
            q = Q()
            for lp in loc_params:
                up = lp.upper()

                # created_by chain (Admin/Manager who created the sale)
                q |= Q(created_by__employee__outlet__location__code__iexact=up)
                q |= Q(created_by__employee__outlet__location__name__icontains=lp)

                # salesman chain (Staff who made the sale)
                q |= Q(salesman__outlet__location__code__iexact=up)
                q |= Q(salesman__outlet__location__name__icontains=lp)

            # Apply filter and use distinct() to avoid duplicate rows from joins
            sales_qs = sales_qs.filter(q).distinct()

        sales = list(sales_qs)

        # ✅ DEBUG: Log what we found
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Day-wise Sales Summary: date_from={df}, date_to={dt}, locations={loc_params}")
        logger.info(f"Found {len(sales)} sales after filtering")

        # If no sales found, return empty result (not 404)
        if not sales:
            return Response({
                "rows": [],
                "totals": {
                    "gross_amount": "0.00",
                    "tax_amount": "0.00",
                    "cgst": "0.00",
                    "sgst": "0.00",
                    "igst": "0.00",
                    "tax_5": "0.00",
                    "tax_18": "0.00",
                    "tax_12": "0.00",
                    "tax_28": "0.00",
                    "cess": "0.00",
                    "discount": "0.00",
                    "bank": "0.00",
                    "cash": "0.00",
                    "credit_notes": "0.00",
                    "coupon_discount": "0.00",
                    "additional_charge": "0.00",
                    "total": "0.00",
                }
            }, status=status.HTTP_200_OK)

        # group by date
        day_map = {}
        for s in sales:
            d = timezone.localtime(s.transaction_date).date()
            day_map.setdefault(d, []).append(s.id)

        # -------- Build rows --------
        rows = []
        totals = {
            "gross_amount": Decimal("0.00"),
            "tax_amount": Decimal("0.00"),
            "cgst": Decimal("0.00"),
            "sgst": Decimal("0.00"),
            "igst": Decimal("0.00"),   # keep for future
            "tax_5": Decimal("0.00"),
            "tax_18": Decimal("0.00"),

            # kept for backward compatibility (UI currently has these cols)
            "tax_12": Decimal("0.00"),
            "tax_28": Decimal("0.00"),
            "cess": Decimal("0.00"),

            "discount": Decimal("0.00"),
            "bank": Decimal("0.00"),
            "cash": Decimal("0.00"),
            "credit_notes": Decimal("0.00"),
            "coupon_discount": Decimal("0.00"),
            "additional_charge": Decimal("0.00"),
            "total": Decimal("0.00"),
        }

        sale_ids_all = [s.id for s in sales]
        lines_all = list(SaleLine.objects.filter(sale_id__in=sale_ids_all).order_by("id")) if sale_ids_all else []
        pays_all = list(SalePayment.objects.filter(sale_id__in=sale_ids_all).order_by("id")) if sale_ids_all else []

        lines_by_sale = {}
        for ln in lines_all:
            lines_by_sale.setdefault(ln.sale_id, []).append(ln)

        pays_by_sale = {}
        for p in pays_all:
            pays_by_sale.setdefault(p.sale_id, []).append(p)

        # credit notes in range + location
        # credit notes in range + location  ✅ use created "date"
        cn_qs = CreditNote.objects.filter(date__range=(start_dt, end_dt))

        if loc_params:
            q = Q()
            for lp in loc_params:
                up = lp.upper()

                # 1) normal notes with location saved
                q |= Q(location__code__iexact=up)
                q |= Q(location__name__icontains=lp)

                # 2) legacy/NULL-location notes: infer location from created_by user
                q |= Q(
                    location__isnull=True,
                    created_by__employee__outlet__location__code__iexact=up,
                )
                q |= Q(
                    location__isnull=True,
                    created_by__employee__outlet__location__name__icontains=lp,
                )

            cn_qs = cn_qs.filter(q).distinct()

        credit_notes = list(cn_qs)
        cn_by_day = {}
        for cn in credit_notes:
            d = timezone.localtime(cn.date).date() if cn.date else None
            if not d:
                continue
            cn_by_day[d] = q2(cn_by_day.get(d, Decimal("0.00")) + q2(cn.amount))

        # iterate days in requested range (so table shows each day)
        cur = df
        sr = 1
        while cur <= dt:
            sale_ids = day_map.get(cur, [])

            gross_amount = Decimal("0.00")  # taxable value sum
            tax_amount = Decimal("0.00")
            cgst = Decimal("0.00")
            sgst = Decimal("0.00")
            igst = Decimal("0.00")  # keep 0 for now

            tax_5 = Decimal("0.00")
            tax_18 = Decimal("0.00")

            # future placeholders (UI safe)
            tax_12 = Decimal("0.00")
            tax_28 = Decimal("0.00")
            cess = Decimal("0.00")

            discount_sum = Decimal("0.00")

            bank = Decimal("0.00")  # UPI + CARD
            cash = Decimal("0.00")  # CASH
            coupon_discount = Decimal("0.00")  # COUPON payment bucket
            additional_charge = Decimal("0.00")  # per requirement

            for sid in sale_ids:
                # --- lines ---
                for ln in lines_by_sale.get(sid, []):
                    qty = Decimal(int(getattr(ln, "qty", 0) or 0))
                    sp = q2(getattr(ln, "sp", 0) or 0)

                    dpu = _per_unit_discount(ln)
                    disc_line = q2(dpu * qty)
                    discount_sum = q2(discount_sum + disc_line)

                    net_unit = q2(sp - dpu)
                    if net_unit < 0:
                        net_unit = Decimal("0.00")

                    inclusive_total = q2(net_unit * qty)

                    rate = _rate_for_unit_amount(net_unit)
                    taxable, tax, c, s = _split_inclusive_amount(inclusive_total, rate)

                    gross_amount = q2(gross_amount + taxable)
                    tax_amount = q2(tax_amount + tax)
                    cgst = q2(cgst + c)
                    sgst = q2(sgst + s)

                    if rate == Decimal("0.05"):
                        tax_5 = q2(tax_5 + tax)
                    else:
                        tax_18 = q2(tax_18 + tax)

                # --- payments ---
                for p in pays_by_sale.get(sid, []):
                    m = (getattr(p, "method", "") or "").upper().strip()
                    amt = q2(getattr(p, "amount", 0) or 0)

                    if m in ("UPI", "CARD"):
                        bank = q2(bank + amt)
                    elif m == "CASH":
                        cash = q2(cash + amt)
                    elif m == "COUPON":
                        coupon_discount = q2(coupon_discount + amt)

            credit_notes_sum = q2(cn_by_day.get(cur, Decimal("0.00")))

            total_amount = q2(gross_amount + tax_amount)

            row = {
                "sr_no": sr,
                "sales_date": cur.strftime("%Y-%m-%d"),
                "gross_amount": f"{gross_amount:.2f}",
                "tax_amount": f"{tax_amount:.2f}",
                "cgst": f"{cgst:.2f}",
                "sgst": f"{sgst:.2f}",
                "igst": f"{igst:.2f}",

                "tax_5": f"{tax_5:.2f}",
                "tax_18": f"{tax_18:.2f}",

                "tax_12": f"{tax_12:.2f}",
                "tax_28": f"{tax_28:.2f}",
                "cess": f"{cess:.2f}",

                "discount": f"{discount_sum:.2f}",
                "bank": f"{bank:.2f}",
                "cash": f"{cash:.2f}",
                "credit_notes": f"{credit_notes_sum:.2f}",
                "coupon_discount": f"{coupon_discount:.2f}",
                "additional_charge": f"{additional_charge:.2f}",
                "total": f"{total_amount:.2f}",
            }

            rows.append(row)

            for k, v in [
                ("gross_amount", gross_amount),
                ("tax_amount", tax_amount),
                ("cgst", cgst),
                ("sgst", sgst),
                ("igst", igst),
                ("tax_5", tax_5),
                ("tax_18", tax_18),
                ("tax_12", tax_12),
                ("tax_28", tax_28),
                ("cess", cess),
                ("discount", discount_sum),
                ("bank", bank),
                ("cash", cash),
                ("credit_notes", credit_notes_sum),
                ("coupon_discount", coupon_discount),
                ("additional_charge", additional_charge),
                ("total", total_amount),
            ]:
                totals[k] = q2(totals[k] + q2(v))

            sr += 1
            cur = cur + timedelta(days=1)

        payload = {
            "rows": rows,
            "totals": {k: f"{q2(v):.2f}" for k, v in totals.items()},
        }

        export = (request.query_params.get("export") or "").strip().lower()
        if export == "pdf":
            return self._export_pdf(df, dt, payload)
        if export in ("xls", "xlsx", "excel"):
            return self._export_excel(df, dt, payload)

        return Response(payload, status=status.HTTP_200_OK)

    def _export_pdf(self, df, dt, payload):
        buf = BytesIO()
        c = canvas.Canvas(buf, pagesize=landscape(A4))
        W, H = landscape(A4)

        title = "Daily Sales Summary (Day Wise)"
        c.setFont("Helvetica-Bold", 14)
        c.drawString(24, H - 30, title)

        c.setFont("Helvetica", 10)
        c.drawString(24, H - 48, f"{df.strftime('%d/%m/%Y')} to {dt.strftime('%d/%m/%Y')}")

        cols = [
            ("Sr", 30),
            ("Sales Date", 70),
            ("Gross", 70),
            ("Tax", 70),
            ("CGST", 60),
            ("SGST", 60),
            ("IGST", 60),
            ("5%", 55),
            ("18%", 55),
            ("Discount", 70),
            ("Bank", 70),
            ("Cash", 70),
            ("Credit Note", 80),
            ("Coupon Disc", 80),
            ("Add. Charge", 80),
            ("Total", 70),
        ]

        x0 = 24
        y = H - 80
        row_h = 16

        def draw_row(values, y, bold=False):
            c.setFont("Helvetica-Bold" if bold else "Helvetica", 9)
            x = x0
            widths = [w for _, w in cols]
            for text, w in zip(values, widths):
                c.drawString(x, y, str(text))
                x += w

        draw_row([name for name, _ in cols], y, bold=True)
        y -= row_h

        for r in payload["rows"]:
            if y < 40:
                c.showPage()
                c.setPageSize(landscape(A4))
                W, H = landscape(A4)
                y = H - 40
                draw_row([name for name, _ in cols], y, bold=True)
                y -= row_h

            draw_row([
                r["sr_no"],
                r["sales_date"],
                r["gross_amount"],
                r["tax_amount"],
                r["cgst"],
                r["sgst"],
                r["igst"],
                r["tax_5"],
                r["tax_18"],
                r["discount"],
                r["bank"],
                r["cash"],
                r["credit_notes"],
                r["coupon_discount"],
                r["additional_charge"],
                r["total"],
            ], y, bold=False)
            y -= row_h

        t = payload["totals"]
        y -= 6
        draw_row([
            "",
            "TOTAL",
            t["gross_amount"],
            t["tax_amount"],
            t["cgst"],
            t["sgst"],
            t["igst"],
            t["tax_5"],
            t["tax_18"],
            t["discount"],
            t["bank"],
            t["cash"],
            t["credit_notes"],
            t["coupon_discount"],
            t["additional_charge"],
            t["total"],
        ], y, bold=True)

        c.save()
        buf.seek(0)

        filename = f"daywise_sales_{df.strftime('%Y%m%d')}_{dt.strftime('%Y%m%d')}.pdf"
        resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    def _export_excel(self, df, dt, payload):
        wb = Workbook()
        ws = wb.active
        ws.title = "Daywise Sales"

        headers = [
            "Sr.No",
            "Sales Date",
            "Gross Amount",
            "Tax Amount",
            "CGST",
            "SGST",
            "IGST",
            "5%",
            "18%",
            "Discount",
            "Bank",
            "Cash",
            "Credit Note",
            "Coupon Discount",
            "Additional Charge",
            "Total",
        ]
        ws.append(headers)

        for r in payload["rows"]:
            ws.append([
                r["sr_no"],
                r["sales_date"],
                r["gross_amount"],
                r["tax_amount"],
                r["cgst"],
                r["sgst"],
                r["igst"],
                r["tax_5"],
                r["tax_18"],
                r["discount"],
                r["bank"],
                r["cash"],
                r["credit_notes"],
                r["coupon_discount"],
                r["additional_charge"],
                r["total"],
            ])

        t = payload["totals"]
        ws.append([
            "",
            "TOTAL",
            t["gross_amount"],
            t["tax_amount"],
            t["cgst"],
            t["sgst"],
            t["igst"],
            t["tax_5"],
            t["tax_18"],
            t["discount"],
            t["bank"],
            t["cash"],
            t["credit_notes"],
            t["coupon_discount"],
            t["additional_charge"],
            t["total"],
        ])

        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        filename = f"daywise_sales_{df.strftime('%Y%m%d')}_{dt.strftime('%Y%m%d')}.xlsx"
        resp = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp
