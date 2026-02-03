# products/views_sales_register.py
from io import BytesIO
from datetime import datetime, time

from django.db.models import Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import RegisterClosing, RegisterSession, CreditNote


def _user_location(user):
    emp = getattr(user, "employee", None)
    outlet = getattr(emp, "outlet", None) if emp else None
    loc = getattr(outlet, "location", None) if outlet else None
    return loc, emp


def _is_admin(user):
    emp = getattr(user, "employee", None)
    role = getattr(emp, "role", "") if emp else ""
    return bool(getattr(user, "is_superuser", False) or role == "ADMIN")


def _ymd(d):
    return d.strftime("%Y-%m-%d")


def _parse_range(request):
    df = parse_date((request.query_params.get("date_from") or "").strip())
    dt = parse_date((request.query_params.get("date_to") or "").strip())
    if not df or not dt:
        return None, None, None, None

    if dt < df:
        df, dt = dt, df

    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime.combine(df, time.min), tz)
    end_dt = timezone.make_aware(datetime.combine(dt, time.max), tz)
    return df, dt, start_dt, end_dt


def _emp_name(emp):
    if not emp:
        return ""
    # best-effort name
    u = getattr(emp, "user", None)
    if u:
        return (u.get_full_name() or u.username or "").strip()
    return (getattr(emp, "name", "") or str(emp)).strip()


class SalesRegisterReportView(APIView):
    """
    GET /api/reports/sales-register/?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD
      optional:
        - location= (repeatable)
        - page, page_size
        - export=pdf | excel

    RULES:
      - Manager: auto restricted to their location; frontend should hide location filter
      - Admin/Superuser: can filter by location(s) and see all if none passed
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        df, dt, start_dt, end_dt = _parse_range(request)
        if not df or not dt:
            return Response(
                {"detail": "date_from and date_to are required (YYYY-MM-DD)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = request.user
        admin = _is_admin(user)
        user_loc, _emp = _user_location(user)

        # locations passed (can repeat)
        loc_params = request.query_params.getlist("location")
        loc_params = [str(x).strip() for x in loc_params if str(x).strip()]

        qs = RegisterClosing.objects.select_related("location", "created_by").filter(
            closed_at__range=(start_dt, end_dt)
        ).order_by("-closed_at", "-id")

        # Manager scoping
        if not admin:
            if user_loc:
                qs = qs.filter(location=user_loc)
            else:
                qs = qs.filter(location__isnull=True)

        # Admin filtering
        if admin and loc_params:
            q = Q()
            for lp in loc_params:
                up = lp.upper()
                # allow filter by code or name
                q |= Q(location__code__iexact=up)
                q |= Q(location__name__icontains=lp)
            qs = qs.filter(q).distinct()

        export = (request.query_params.get("export") or "").strip().lower()

        # pagination (ignored when exporting)
        try:
            page_size = int(request.query_params.get("page_size", 10))
        except Exception:
            page_size = 10
        page_size = max(1, min(1000, page_size))

        try:
            page = int(request.query_params.get("page", 1))
        except Exception:
            page = 1
        page = max(1, page)

        total = qs.count()

        if export in ("pdf", "excel", "xlsx", "xls"):
            closings = list(qs)
            rows, totals = self._build_rows(closings)
            if export == "pdf":
                return self._export_pdf(df, dt, rows, totals, admin)
            return self._export_excel(df, dt, rows, totals, admin)

        start_ix = (page - 1) * page_size
        end_ix = start_ix + page_size
        closings = list(qs[start_ix:end_ix])

        rows, totals = self._build_rows(closings)

        return Response(
            {
                "results": rows,
                "total": total,
                "page": page,
                "page_size": page_size,
                "totals": totals,
            },
            status=status.HTTP_200_OK,
        )

    def _build_rows(self, closings):
        tz = timezone.get_current_timezone()

        # collect (location_id, business_date) for opening cash lookups
        keys = []
        for rc in closings:
            d = timezone.localtime(rc.closed_at, tz).date()
            keys.append((rc.location_id, d))

        # bulk sessions
        sessions = RegisterSession.objects.filter(
            Q(location_id__in=[k[0] for k in keys if k[0] is not None])
        ).values("location_id", "business_date", "opening_cash")
        sess_map = {(x["location_id"], x["business_date"]): x["opening_cash"] for x in sessions}

        # bulk credit notes made sums per (location, date)
        # credit note "made that day" -> use CreditNote.date
        cn_q = CreditNote.objects.all()
        loc_ids = list({rc.location_id for rc in closings if rc.location_id})
        if loc_ids:
            cn_q = cn_q.filter(location_id__in=loc_ids)

        # date window for CNs = min/max of rows (safe)
        if closings:
            min_dt = min([rc.closed_at for rc in closings])
            max_dt = max([rc.closed_at for rc in closings])
            cn_q = cn_q.filter(date__range=(min_dt.replace(hour=0, minute=0, second=0, microsecond=0),
                                            max_dt.replace(hour=23, minute=59, second=59, microsecond=999999)))

        # build cn sum map per (location_id, date)
        cn_map = {}
        for cn in cn_q.values("location_id", "date").annotate(total=Sum("amount")):
            loc_id = cn["location_id"]
            day = timezone.localtime(cn["date"], tz).date() if cn["date"] else None
            if not day:
                continue
            cn_map[(loc_id, day)] = (cn_map.get((loc_id, day), 0) or 0) + (cn["total"] or 0)

        rows = []
        # totals for footer
        t_open = 0
        t_cash = 0
        t_card = 0
        t_upi = 0
        t_total_sales = 0
        t_credit_applied = 0
        t_sales_return = 0
        t_closing = 0

        for i, rc in enumerate(closings, start=1):
            d = timezone.localtime(rc.closed_at, tz).date()
            loc = getattr(rc, "location", None)
            loc_name = (getattr(loc, "name", "") or "").strip()
            loc_code = (getattr(loc, "code", "") or "").strip()
            location_label = loc_name or loc_code or ""

            opening_cash = sess_map.get((rc.location_id, d), getattr(rc, "opening_cash", 0) or 0)

            cash_amt = getattr(rc, "cash_payment", 0) or 0
            card_amt = getattr(rc, "card_payment", 0) or 0
            upi_amt = getattr(rc, "upi_payment", 0) or 0

            total_sales = getattr(rc, "total_sales", 0) or 0
            credit_applied = getattr(rc, "credit_applied", 0) or 0

            # sales return amount = credit notes MADE that day
            sales_return_amt = cn_map.get((rc.location_id, d), 0) or 0

            # closing amount from close register
                        # ✅ Closing Amount = manager entered "Total Cash Left In Drawer"
            closing_amt = getattr(rc, "closing_amount", None)
            if closing_amt is None:
                closing_amt = getattr(rc, "total_cash_left", 0) or 0


            cashier_name = _emp_name(getattr(rc, "created_by", None))

            rows.append(
                {
                    "sr": i,
                    "location": location_label,
                    "cashier": cashier_name,
                    "from_date": _ymd(d),
                    "to_date": _ymd(d),
                    "cash_in_hand": f"{float(opening_cash):.2f}",
                    "cash": f"{float(cash_amt):.2f}",
                    "card": f"{float(card_amt):.2f}",
                    "upi": f"{float(upi_amt):.2f}",
                    "total_sales": f"{float(total_sales):.2f}",
                    "credit_applied_amount": f"{float(credit_applied):.2f}",
                    "sales_return_amount": f"{float(sales_return_amt):.2f}",
                    "closing_amount": f"{float(closing_amt):.2f}",
                }
            )

            t_open += float(opening_cash or 0)
            t_cash += float(cash_amt or 0)
            t_card += float(card_amt or 0)
            t_upi += float(upi_amt or 0)
            t_total_sales += float(total_sales or 0)
            t_credit_applied += float(credit_applied or 0)
            t_sales_return += float(sales_return_amt or 0)
            t_closing += float(closing_amt or 0)

        totals = {
            "cash_in_hand": f"{t_open:.2f}",
            "cash": f"{t_cash:.2f}",
            "card": f"{t_card:.2f}",
            "upi": f"{t_upi:.2f}",
            "total_sales": f"{t_total_sales:.2f}",
            "credit_applied_amount": f"{t_credit_applied:.2f}",
            "sales_return_amount": f"{t_sales_return:.2f}",
            "closing_amount": f"{t_closing:.2f}",
        }
        return rows, totals

    def _export_pdf(self, df, dt, rows, totals, admin):
        buf = BytesIO()
        c = canvas.Canvas(buf, pagesize=landscape(A4))
        W, H = landscape(A4)

        c.setFont("Helvetica-Bold", 14)
        c.drawString(24, H - 30, "Sales Register")

        c.setFont("Helvetica", 10)
        c.drawString(24, H - 48, f"{df.strftime('%d/%m/%Y')} to {dt.strftime('%d/%m/%Y')}")

        # columns
        cols = []
        if admin:
            cols.append(("Location", 140))
        cols += [
            ("#", 28),
            ("Cashier", 140),
            ("From", 70),
            ("To", 70),
            ("Cash In Hand", 85),
            ("Cash", 70),
            ("Card", 70),
            ("UPI", 70),
            ("Total Sales", 85),
            ("Credit Applied", 95),
            ("Sales Return", 90),
            ("Closing Amount", 95),
        ]

        x0 = 24
        y = H - 80
        row_h = 16

        def draw_row(values, y, bold=False):
            c.setFont("Helvetica-Bold" if bold else "Helvetica", 9)
            x = x0
            for v, (_, w) in zip(values, cols):
                c.drawString(x, y, str(v))
                x += w

        draw_row([name for name, _ in cols], y, bold=True)
        y -= row_h

        for r in rows:
            if y < 40:
                c.showPage()
                c.setPageSize(landscape(A4))
                W, H = landscape(A4)
                y = H - 40
                draw_row([name for name, _ in cols], y, bold=True)
                y -= row_h

            values = []
            if admin:
                values.append(r.get("location", ""))
            values += [
                r.get("sr", ""),
                r.get("cashier", ""),
                r.get("from_date", ""),
                r.get("to_date", ""),
                r.get("cash_in_hand", ""),
                r.get("cash", ""),
                r.get("card", ""),
                r.get("upi", ""),
                r.get("total_sales", ""),
                r.get("credit_applied_amount", ""),
                r.get("sales_return_amount", ""),
                r.get("closing_amount", ""),
            ]
            draw_row(values, y, bold=False)
            y -= row_h

        # totals row
        y -= 8
        values = []
        if admin:
            values.append("")
        values += [
            "",
            "TOTAL",
            "",
            "",
            totals["cash_in_hand"],
            totals["cash"],
            totals["card"],
            totals["upi"],
            totals["total_sales"],
            totals["credit_applied_amount"],
            totals["sales_return_amount"],
            totals["closing_amount"],
        ]
        draw_row(values, y, bold=True)

        c.save()
        buf.seek(0)

        filename = f"sales_register_{df.strftime('%Y%m%d')}_{dt.strftime('%Y%m%d')}.pdf"
        resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    def _export_excel(self, df, dt, rows, totals, admin):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Register"

        headers = []
        if admin:
            headers.append("Location")
        headers += [
            "#",
            "Cashier",
            "From Date",
            "To Date",
            "Cash In Hand",
            "Cash",
            "Card",
            "UPI",
            "Total Sales",
            "Credit Applied Amount",
            "Sales Return Amount",
            "Closing Amount",
        ]
        ws.append(headers)

        for r in rows:
            row = []
            if admin:
                row.append(r.get("location", ""))
            row += [
                r.get("sr", ""),
                r.get("cashier", ""),
                r.get("from_date", ""),
                r.get("to_date", ""),
                r.get("cash_in_hand", ""),
                r.get("cash", ""),
                r.get("card", ""),
                r.get("upi", ""),
                r.get("total_sales", ""),
                r.get("credit_applied_amount", ""),
                r.get("sales_return_amount", ""),
                r.get("closing_amount", ""),
            ]
            ws.append(row)

        # totals
        tr = []
        if admin:
            tr.append("")
        tr += [
            "",
            "TOTAL",
            "",
            "",
            totals["cash_in_hand"],
            totals["cash"],
            totals["card"],
            totals["upi"],
            totals["total_sales"],
            totals["credit_applied_amount"],
            totals["sales_return_amount"],
            totals["closing_amount"],
        ]
        ws.append(tr)

        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(h)) + 2)

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        filename = f"sales_register_{df.strftime('%Y%m%d')}_{dt.strftime('%Y%m%d')}.xlsx"
        resp = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp
