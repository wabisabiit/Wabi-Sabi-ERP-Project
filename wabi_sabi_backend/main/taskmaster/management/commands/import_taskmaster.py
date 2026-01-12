# taskmaster/management/commands/import_taskmaster.py

import os
import re
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from django.db import transaction

from taskmaster.models import TaskItem

try:
    import openpyxl
except ImportError:
    openpyxl = None


def _norm_header(h):
    return (
        str(h or "")
        .strip()
        .lower()
        .replace("\n", " ")
        .replace("\t", " ")
        .replace(".", "")
        .replace("-", "_")
    )


def _clean_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _to_bool(v, default=True):
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("1", "true", "yes", "y", "active"):
        return True
    if s in ("0", "false", "no", "n", "inactive"):
        return False
    return default


def _to_decimal(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        return None


class Command(BaseCommand):
    help = "Import/Update TaskItem rows from an Excel file kept inside management/commands folder."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            dest="file",
            default="",
            help="Optional path to xlsx. If not provided, command will auto-pick the first .xlsx in this folder.",
        )
        parser.add_argument(
            "--sheet",
            dest="sheet",
            default="",
            help="Optional sheet name. If not provided, first sheet is used.",
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            self.stderr.write("❌ openpyxl not installed. Run: pip install openpyxl")
            return

        cmd_dir = os.path.dirname(os.path.abspath(__file__))

        xlsx_path = (options.get("file") or "").strip()
        if not xlsx_path:
            cands = [
                os.path.join(cmd_dir, f)
                for f in os.listdir(cmd_dir)
                if f.lower().endswith(".xlsx") and not f.startswith("~$")
            ]
            if not cands:
                self.stderr.write(f"❌ No .xlsx found in: {cmd_dir}")
                return
            cands.sort(key=lambda p: os.path.getmtime(p), reverse=True)
            xlsx_path = cands[0]

        if not os.path.isabs(xlsx_path):
            xlsx_path = os.path.join(os.getcwd(), xlsx_path)

        if not os.path.exists(xlsx_path):
            self.stderr.write(f"❌ Excel file not found: {xlsx_path}")
            return

        self.stdout.write(f"Loading Excel: {xlsx_path}")

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        sheet_name = (options.get("sheet") or "").strip()
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                self.stderr.write(f"❌ Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
                return
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        # Header row
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        headers_raw = [c.value for c in header_row]
        headers = [_norm_header(h) for h in headers_raw]

        def idx(*names):
            norms = [_norm_header(n) for n in names]
            for n in norms:
                if n in headers:
                    return headers.index(n)
            return -1

        # ✅ Your file uses these headers:
        # ['Code', 'Department', 'Category', 'Item Name', 'Print Name', 'Hsn Code']
        i_item_code = idx("code", "item_code", "item code", "itemcode", "sku")
        if i_item_code < 0:
            self.stderr.write("❌ Could not find Item Code column. Header row is:")
            self.stderr.write(str(headers_raw))
            return

        i_department = idx("department")
        i_category = idx("category")
        i_full = idx("item name", "item_name", "item full name", "item_full_name", "name")
        i_print = idx("print name", "print_name", "item_print_friendly_name")
        i_hsn = idx("hsn code", "hsn_code", "hsn")

        # Optional (not in your current excel, but safe if later added)
        i_vasy = idx("item vasy name", "item_vasy_name", "vasy name")
        i_active = idx("item_active", "active", "is_active", "status")
        i_gst = idx("gst", "gst%","gst %","gst rate")

        created = 0
        updated = 0
        skipped = 0
        errors = 0

        known_idxs = {
            i for i in [i_item_code, i_category, i_department, i_full, i_vasy, i_print, i_active, i_hsn, i_gst]
            if i >= 0
        }

        @transaction.atomic
        def run_import():
            nonlocal created, updated, skipped, errors

            for row_i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                values = [c.value for c in row]

                raw_code = values[i_item_code] if i_item_code < len(values) else None
                item_code = _clean_spaces(str(raw_code or ""))
                if not item_code:
                    skipped += 1
                    continue

                # keep formats like "100" or "100-W"
                item_code = item_code.upper()

                category = _clean_spaces(str(values[i_category] or "")) if (i_category >= 0 and i_category < len(values)) else ""
                department = _clean_spaces(str(values[i_department] or "")) if (i_department >= 0 and i_department < len(values)) else ""

                item_full_name = _clean_spaces(str(values[i_full] or "")) if (i_full >= 0 and i_full < len(values)) else ""
                item_print_friendly_name = _clean_spaces(str(values[i_print] or "")) if (i_print >= 0 and i_print < len(values)) else ""
                item_vasy_name = _clean_spaces(str(values[i_vasy] or "")) if (i_vasy >= 0 and i_vasy < len(values)) else ""

                hsn_code = _clean_spaces(str(values[i_hsn] or "")) if (i_hsn >= 0 and i_hsn < len(values)) else ""

                item_active = _to_bool(values[i_active], default=True) if (i_active >= 0 and i_active < len(values)) else True
                gst = _to_decimal(values[i_gst]) if (i_gst >= 0 and i_gst < len(values)) else None

                # store extra columns into attributes (if any)
                attrs = {}
                for col_idx, h in enumerate(headers):
                    if col_idx in known_idxs:
                        continue
                    key = _clean_spaces(h).replace(" ", "_")
                    if not key:
                        continue
                    val = values[col_idx] if col_idx < len(values) else None
                    if val is None or str(val).strip() == "":
                        continue
                    try:
                        attrs[key] = val
                    except Exception:
                        attrs[key] = str(val)

                try:
                    obj = TaskItem.objects.filter(item_code=item_code).first()
                    if not obj:
                        TaskItem.objects.create(
                            item_code=item_code,
                            category=category,
                            department=department,
                            item_full_name=item_full_name,
                            item_vasy_name=item_vasy_name,
                            item_print_friendly_name=item_print_friendly_name,
                            item_active=item_active,
                            hsn_code=hsn_code,
                            gst=gst,
                            attributes=attrs or {},
                        )
                        created += 1
                    else:
                        changed = False

                        def set_if(field, value):
                            nonlocal changed
                            if getattr(obj, field) != value:
                                setattr(obj, field, value)
                                changed = True

                        set_if("category", category)
                        set_if("department", department)
                        set_if("item_full_name", item_full_name)
                        set_if("item_vasy_name", item_vasy_name)
                        set_if("item_print_friendly_name", item_print_friendly_name)
                        set_if("item_active", item_active)
                        set_if("hsn_code", hsn_code)

                        if obj.gst != gst:
                            obj.gst = gst
                            changed = True

                        if attrs:
                            merged = dict(obj.attributes or {})
                            for k, v in attrs.items():
                                if merged.get(k) != v:
                                    merged[k] = v
                                    changed = True
                            obj.attributes = merged

                        if changed:
                            obj.save()
                            updated += 1
                        else:
                            skipped += 1

                except Exception as e:
                    errors += 1
                    self.stderr.write(f"❌ Row {row_i} failed for item_code={item_code}: {e}")

        run_import()

        self.stdout.write(self.style.SUCCESS("✅ TaskMaster import completed"))
        self.stdout.write(f"Created: {created}")
        self.stdout.write(f"Updated: {updated}")
        self.stdout.write(f"Skipped (no change/blank rows): {skipped}")
        self.stdout.write(f"Errors: {errors}")
